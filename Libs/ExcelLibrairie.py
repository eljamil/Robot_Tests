import os
import openpyxl


class ExcelLibrairie(object):
    @staticmethod
    def read_excel_sheet(filename, sheet_name):
        """This creates a keyword named "Read Excel Sheet"

        This keyword takes two arguments:
        - filename: path to the Excel (.xlsx) file.
        - sheet_name: name of the sheet to read from.

        It returns a list of rows, with each row being a list of the data in
        each column.
        """
        data = []
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=1, values_only=True):  # Start from the second row
                if any(row):  # Check if the row contains any data
                    data.append(row)
            return data
        except FileNotFoundError:
            print(f"File '{filename}' not found.")
            return None
        except KeyError:
            print(f"Sheet '{sheet_name}' not found in '{filename}'.")
            return None

    @staticmethod
    def read_excel_row(filename, sheet_name, row_number):
        """
            Reads a specific row from the Excel sheet.
            :param filename: path to the Excel (.xlsx) file.
            :param sheet_name: name of the sheet to read from.
            :param row_number: the row number to read (1-based indexing).
            :return: a list containing the data of the specified row, or None if the row is not found.
        """
        data = ExcelLibrairie.read_excel_sheet(filename, sheet_name)
        if data:
            try:
                return data[row_number - 1]
            except IndexError:
                print(f"Row number {row_number} not found in '{sheet_name}' of '{filename}'.")
                return None
        else:
            return None

    @staticmethod
    def read_excel_first_row(filename, sheet_name):
        """
        Reads the first row from the Excel sheet.

        :param filename: path to the Excel (.xlsx) file.
        :param sheet_name: name of the sheet to read from.
        :return: a list containing the data of the first row, or None if the sheet is empty.
        """
        data = ExcelLibrairie.read_excel_sheet(filename, sheet_name)
        if data:
            return data[0]
        else:
            return None

    @staticmethod
    def read_env_row_a(filename, sheet_name, env):
        """
        Reads the row corresponding to the given environment from the Excel sheet.

        :param filename: path to the Excel (.xlsx) file.
        :param sheet_name: name of the sheet to read from.
        :param env: the environment for which the row is to be retrieved.
        :return: a list containing the data of the row corresponding to the given environment,
                 or None if the environment is not found.
        """
        data = ExcelLibrairie.read_excel_sheet(filename, sheet_name)
        if data:
            for row in data:
                if row[0] == env:
                    return row
            print(f"Environment '{env}' not found in '{sheet_name}' of '{filename}'.")
            return None
        else:
            return None

    @staticmethod
    def read_data(excel_file_path, sheet_name):
        """
        Reads data from an Excel sheet located at the specified path
        and returns it as a dictionary.

        :param excel_file_path: full path to the Excel file.
        :param sheet_name: name of the sheet to read from.
        :return: a dictionary containing the data read from the Excel sheet,
                 with "${column_name}" as keys and row values as values.
        """
        data = ExcelLibrairie.read_excel_sheet(excel_file_path, sheet_name)
        if data:
            # Assuming the first row contains the column headers
            headers = data[0]
            # Construct the dictionary with all columns
            data_dict = {}
            for row in data[1:]:
                row_dict = {header: value for header, value in zip(headers, row)}
                data_dict.update(row_dict)
            return data_dict
        else:
            return None

    @staticmethod
    def write_data2(excel_file_path, sheet_name, column_name, new_value):
        """
        Writes a new value to all cells in a specific column in an Excel sheet located at the specified path.

        :param excel_file_path: full path to the Excel file.
        :param sheet_name: name of the sheet to write to.
        :param column_name: the name of the column to update.
        :param new_value: the new value to write to the column.
        """
        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb[sheet_name]

        # Find the column index for the column name
        column_index = None
        for col_idx, cell in enumerate(sheet[1]):
            print(f"Checking cell {cell.value} in column {col_idx+1}")
            if cell.value == column_name:
                column_index = col_idx + 1  # openpyxl uses 1-based indexing
                break

        if column_index is None:
            raise ValueError(f"Column {column_name} not found in sheet {sheet_name}")

        # Write the new value to all cells in the column (except the header)
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=column_index, value=new_value)

        wb.save(excel_file_path)

    @staticmethod
    def write_data(excel_file_path, sheet_name, column_name, new_value):
        """
        Écrit une nouvelle valeur dans la première cellule d'une colonne spécifique (après l'en-tête) dans une feuille Excel.

        :param excel_file_path: chemin complet du fichier Excel.
        :param sheet_name: nom de la feuille à écrire.
        :param column_name: le nom de la colonne à mettre à jour.
        :param new_value: la nouvelle valeur à écrire dans la cellule.
        """
        # Charger le fichier Excel
        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb[sheet_name]

        # Trouver l'index de la colonne pour le nom de colonne spécifié
        column_index = None
        for col_idx, cell in enumerate(sheet[1]):
            if cell.value == column_name:
                column_index = col_idx + 1  # openpyxl utilise une indexation à partir de 1
                break

        # Si la colonne n'est pas trouvée, lever une exception
        if column_index is None:
            raise ValueError(f"La colonne {column_name} n'a pas été trouvée dans la feuille {sheet_name}")

        # Écrire la nouvelle valeur dans la première cellule de la colonne (après l'en-tête)
        first_data_row = 2  # Commence après l'en-tête (ligne 1)
        sheet.cell(row=first_data_row, column=column_index, value=new_value)

        # Sauvegarder le fichier Excel
        wb.save(excel_file_path)

    @staticmethod
    def read_data1(filename, sheet_name='NONE'):
        """This function reads an entire Excel file and converts the data to a structured dictionary.

        This keyword takes two arguments:
        - filename: path to the Excel (.xlsx) file.
        - sheet_name: name of the sheet to read from. Default is 'NONE' to read all sheets.

        It returns a dictionary with sheet names as keys and another dictionary as values,
        where each key is a column name and each value is the corresponding data.
        """
        if sheet_name == 'NONE':
            # Read all sheets into a single dictionary
            result = {}
            try:
                workbook = openpyxl.load_workbook(filename)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet_data = []
                    for row in sheet.iter_rows(min_row=1, values_only=True):
                        if any(row):  # Check if the row contains any data
                            sheet_data.append(row)

                    if sheet_data:
                        headers = sheet_data[0]
                        sheet_dict = {}
                        for row in sheet_data[1:]:
                            row_dict = {header: value for header, value in zip(headers, row)}
                            sheet_dict.update(row_dict)  # Update sheet_dict directly with row_dict
                        result[sheet_name] = sheet_dict
                return result
            except FileNotFoundError:
                print(f"File '{filename}' not found.")
                return None
            except Exception as e:
                print(f"An error occurred: {str(e)}")
                return None
        else:
            # Read specific sheet and return a flattened dictionary
            data = ExcelLibrairie.read_excel_sheet(filename, sheet_name)
            if data:
                # Assuming the first row contains the column headers
                headers = data[0]
                # Construct the dictionary with all columns
                data_dict = {}
                for row in data[1:]:
                    row_dict = {header: value for header, value in zip(headers, row)}
                    data_dict.update(row_dict)
                return data_dict
            else:
                return None

    @staticmethod
    def read_first_data_row(filename, sheet_name):
        """
        Reads the first data row (after the header) from the Excel sheet and returns it as a dictionary.

        :param filename: path to the Excel (.xlsx) file.
        :param sheet_name: name of the sheet to read from.
        :return: a dictionary containing the data of the first row, with column headers as keys,
                 or None if the sheet is empty.
        """
        data = ExcelLibrairie.read_excel_sheet(filename, sheet_name)
        if data and len(data) > 1:
            # Assuming the first row contains the column headers
            headers = data[0]
            # Take the first data row after the headers
            first_row = data[1]
            # Return a dictionary mapping column headers to the values of the first data row
            return {header: value for header, value in zip(headers, first_row)}
        else:
            print(f"No data found in '{sheet_name}' of '{filename}'.")
            return None